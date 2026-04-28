from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import AuthCredential, User, UserRole
from app.services.audit import write_audit_log
from app.services.auth import hash_password

DEFAULT_REVIEWER_PASSWORD = "reviewer123"
DEFAULT_STUDENT_PASSWORD = "student123"

REVIEWER_HEADERS = ["姓名", "用户名", "邮箱", "科室"]
STUDENT_HEADERS = ["姓名", "用户名", "学号", "邮箱"]


def build_template_workbook_bytes(headers: list[str], sample_rows: list[list[str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"
    sheet.append(headers)
    for row in sample_rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def reviewer_template_bytes() -> bytes:
    return build_template_workbook_bytes(
        REVIEWER_HEADERS,
        [
            ["张老师", "reviewer_zhang", "zhang@example.com", "计算机系"],
            ["李老师", "reviewer_li", "li@example.com", "软件系"],
        ],
    )


def student_template_bytes() -> bytes:
    return build_template_workbook_bytes(
        STUDENT_HEADERS,
        [
            ["王同学", "student_wang", "20260001", "wang@example.com"],
            ["赵同学", "student_zhao", "20260002", "zhao@example.com"],
        ],
    )


def import_reviewers_from_excel(
    db: Session,
    actor_id: int,
    content: bytes,
    default_password: str | None = None,
) -> dict:
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    _validate_headers(sheet, REVIEWER_HEADERS)
    password_to_use = _resolve_default_password(default_password, DEFAULT_REVIEWER_PASSWORD)

    existing_usernames = {
        username.lower() for username in db.scalars(select(AuthCredential.username)).all()
    }
    batch_usernames: set[str] = set()
    created_count = 0
    failures: list[dict] = []

    for row_index, values in _iter_sheet_rows(sheet, len(REVIEWER_HEADERS)):
        name, username, email, department = values
        normalized_username = username.lower()

        reason = _validate_common_account_row(name=name, username=username)
        if not reason and normalized_username in existing_usernames:
            reason = "用户名已存在"
        if not reason and normalized_username in batch_usernames:
            reason = "表格内用户名重复"

        if reason:
            failures.append({"row": row_index, "reason": reason})
            continue

        reviewer = User(
            role=UserRole.REVIEWER,
            name=name,
            email=email or None,
            department=department or None,
        )
        db.add(reviewer)
        db.flush()
        db.add(
            AuthCredential(
                user_id=reviewer.id,
                username=username,
                password_hash=hash_password(password_to_use),
                is_active=True,
            )
        )
        write_audit_log(
            db,
            actor_id,
            "reviewer_import_create",
            "user",
            str(reviewer.id),
            payload={"username": username},
        )
        existing_usernames.add(normalized_username)
        batch_usernames.add(normalized_username)
        created_count += 1

    db.commit()
    return {
        "created_count": created_count,
        "failed_count": len(failures),
        "failures": failures,
        "default_password": password_to_use,
    }


def import_students_from_excel(
    db: Session,
    actor_id: int,
    content: bytes,
    default_password: str | None = None,
) -> dict:
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    _validate_headers(sheet, STUDENT_HEADERS)
    password_to_use = _resolve_default_password(default_password, DEFAULT_STUDENT_PASSWORD)

    existing_usernames = {
        username.lower() for username in db.scalars(select(AuthCredential.username)).all()
    }
    existing_student_nos = {
        student_no.strip().lower()
        for student_no in db.scalars(select(User.student_no).where(User.student_no.is_not(None))).all()
        if student_no
    }
    batch_usernames: set[str] = set()
    batch_student_nos: set[str] = set()
    created_count = 0
    failures: list[dict] = []

    for row_index, values in _iter_sheet_rows(sheet, len(STUDENT_HEADERS)):
        name, username, student_no, email = values
        normalized_username = username.lower()
        normalized_student_no = student_no.lower() if student_no else ""

        reason = _validate_common_account_row(name=name, username=username)
        if not reason and not student_no:
            reason = "学号不能为空"
        if not reason and normalized_username in existing_usernames:
            reason = "用户名已存在"
        if not reason and normalized_username in batch_usernames:
            reason = "表格内用户名重复"
        if not reason and normalized_student_no in existing_student_nos:
            reason = "学号已存在"
        if not reason and normalized_student_no in batch_student_nos:
            reason = "表格内学号重复"

        if reason:
            failures.append({"row": row_index, "reason": reason})
            continue

        student = User(
            role=UserRole.STUDENT,
            name=name,
            student_no=student_no,
            email=email or None,
        )
        db.add(student)
        db.flush()
        db.add(
            AuthCredential(
                user_id=student.id,
                username=username,
                password_hash=hash_password(password_to_use),
                is_active=True,
            )
        )
        write_audit_log(
            db,
            actor_id,
            "student_import_create",
            "user",
            str(student.id),
            payload={"username": username, "student_no": student_no},
        )
        existing_usernames.add(normalized_username)
        existing_student_nos.add(normalized_student_no)
        batch_usernames.add(normalized_username)
        batch_student_nos.add(normalized_student_no)
        created_count += 1

    db.commit()
    return {
        "created_count": created_count,
        "failed_count": len(failures),
        "failures": failures,
        "default_password": password_to_use,
    }


def _validate_headers(sheet, expected_headers: list[str]) -> None:
    header_values = [sheet.cell(row=1, column=index + 1).value for index in range(len(expected_headers))]
    if header_values != expected_headers:
        raise ValueError("导入模板表头不正确，请重新下载模板。")


def _iter_sheet_rows(sheet, column_count: int) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=column_count, values_only=True),
        start=2,
    ):
        values = [_normalize_excel_cell(value) for value in row]
        if not any(values):
            continue
        rows.append((row_index, values))
    return rows


def _normalize_excel_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _validate_common_account_row(*, name: str, username: str) -> str | None:
    if not name:
        return "姓名不能为空"
    if not username:
        return "用户名不能为空"
    if len(username) < 3:
        return "用户名至少 3 个字符"
    return None


def _resolve_default_password(custom_password: str | None, fallback_password: str) -> str:
    password = (custom_password or "").strip()
    if not password:
        return fallback_password
    if len(password) < 6:
        raise ValueError("默认初始密码至少 6 个字符。")
    return password
